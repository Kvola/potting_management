# -*- coding: utf-8 -*-

import base64
import io
from datetime import datetime

from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class PottingImportContractsWizard(models.TransientModel):
    """Wizard pour importer des contrats (commandes clients) depuis un fichier Excel."""
    
    _name = 'potting.import.contracts.wizard'
    _description = 'Import de contrats depuis Excel'

    # -------------------------------------------------------------------------
    # FIELDS
    # -------------------------------------------------------------------------
    file_data = fields.Binary(
        string="Fichier Excel",
        help="Fichier Excel (.xlsx) contenant les contrats à importer"
    )
    
    file_name = fields.Char(
        string="Nom du fichier"
    )
    
    template_file = fields.Binary(
        string="Modèle Excel",
        compute='_compute_template_file'
    )
    
    template_file_name = fields.Char(
        string="Nom du modèle",
        compute='_compute_template_file'
    )
    
    state = fields.Selection([
        ('upload', 'Téléchargement'),
        ('preview', 'Aperçu'),
        ('done', 'Terminé'),
    ], string="État", default='upload')
    
    preview_line_ids = fields.One2many(
        'potting.import.contracts.wizard.line',
        'wizard_id',
        string="Lignes à importer"
    )
    
    total_lines = fields.Integer(
        string="Total lignes",
        compute='_compute_stats'
    )
    
    valid_lines = fields.Integer(
        string="Lignes valides",
        compute='_compute_stats'
    )
    
    error_lines = fields.Integer(
        string="Lignes en erreur",
        compute='_compute_stats'
    )
    
    import_result = fields.Text(
        string="Résultat de l'import"
    )

    # -------------------------------------------------------------------------
    # COMPUTE METHODS
    # -------------------------------------------------------------------------
    @api.depends('preview_line_ids', 'preview_line_ids.is_valid')
    def _compute_stats(self):
        for wizard in self:
            wizard.total_lines = len(wizard.preview_line_ids)
            wizard.valid_lines = len(wizard.preview_line_ids.filtered('is_valid'))
            wizard.error_lines = wizard.total_lines - wizard.valid_lines

    def _compute_template_file(self):
        """Generate the Excel template file."""
        for wizard in self:
            if not OPENPYXL_AVAILABLE:
                wizard.template_file = False
                wizard.template_file_name = False
                continue
            
            template_data = self._generate_template()
            wizard.template_file = base64.b64encode(template_data)
            wizard.template_file_name = 'modele_import_contrats.xlsx'

    # -------------------------------------------------------------------------
    # TEMPLATE GENERATION
    # -------------------------------------------------------------------------
    def _generate_template(self):
        """Generate an Excel template with example data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contrats"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Headers
        headers = [
            ("A", "Numéro de contrat *", 20),
            ("B", "Client (Nom) *", 25),
            ("C", "Type de produit *", 20),
            ("D", "Tonnage contrat (T) *", 18),
            ("E", "Prix unitaire", 15),
            ("F", "Devise", 10),
            ("G", "Date de commande", 15),
            ("H", "Date livraison prévue", 18),
            ("I", "Campagne", 15),
            ("J", "Certifications", 25),
            ("K", "Taux droits export (%)", 20),
            ("L", "Notes", 30),
        ]
        
        for col, header_name, width in headers:
            cell = ws[f"{col}1"]
            cell.value = header_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[col].width = width
        
        # Example data rows
        examples = [
            {
                'contract_number': 'CONT-2025-001',
                'customer': 'BARRY CALLEBAUT',
                'product_type': 'cocoa_mass',
                'tonnage': 500.0,
                'unit_price': 2500,
                'currency': 'EUR',
                'date_order': '2025-01-15',
                'date_expected': '2025-03-15',
                'campaign': '2024-2025',
                'certifications': 'UTZ, Rainforest Alliance',
                'export_duty_rate': 14.6,
                'notes': 'Livraison port d\'Abidjan',
            },
            {
                'contract_number': 'CONT-2025-002',
                'customer': 'CARGILL',
                'product_type': 'cocoa_butter',
                'tonnage': 250.5,
                'unit_price': 3200,
                'currency': 'EUR',
                'date_order': '2025-01-20',
                'date_expected': '2025-04-01',
                'campaign': '2024-2025',
                'certifications': 'Fairtrade',
                'export_duty_rate': 14.6,
                'notes': '',
            },
            {
                'contract_number': 'CONT-2025-003',
                'customer': 'OLAM',
                'product_type': 'cocoa_powder',
                'tonnage': 100.0,
                'unit_price': 1800,
                'currency': 'EUR',
                'date_order': '2025-02-01',
                'date_expected': '2025-05-15',
                'campaign': '2024-2025',
                'certifications': '',
                'export_duty_rate': 14.6,
                'notes': 'Qualité premium',
            },
        ]
        
        for row_idx, example in enumerate(examples, start=2):
            ws[f"A{row_idx}"] = example['contract_number']
            ws[f"B{row_idx}"] = example['customer']
            ws[f"C{row_idx}"] = example['product_type']
            ws[f"D{row_idx}"] = example['tonnage']
            ws[f"E{row_idx}"] = example['unit_price']
            ws[f"F{row_idx}"] = example['currency']
            ws[f"G{row_idx}"] = example['date_order']
            ws[f"H{row_idx}"] = example['date_expected']
            ws[f"I{row_idx}"] = example['campaign']
            ws[f"J{row_idx}"] = example['certifications']
            ws[f"K{row_idx}"] = example['export_duty_rate']
            ws[f"L{row_idx}"] = example['notes']
            
            # Apply example style
            for col in "ABCDEFGHIJKL":
                cell = ws[f"{col}{row_idx}"]
                cell.fill = example_fill
                cell.border = thin_border
        
        # Instructions sheet
        ws_help = wb.create_sheet("Instructions")
        ws_help.column_dimensions['A'].width = 25
        ws_help.column_dimensions['B'].width = 60
        
        instructions = [
            ("Colonne", "Description"),
            ("Numéro de contrat *", "Numéro unique du contrat (obligatoire)"),
            ("Client (Nom) *", "Nom exact du client dans Odoo (obligatoire)"),
            ("Type de produit *", "cocoa_mass, cocoa_butter, cocoa_cake ou cocoa_powder (obligatoire)"),
            ("Tonnage (T) *", "Tonnage en tonnes (nombre décimal, obligatoire)"),
            ("Prix unitaire", "Prix par tonne (optionnel)"),
            ("Devise", "Code devise: EUR, XOF, USD (défaut: EUR)"),
            ("Date de commande", "Format: AAAA-MM-JJ (défaut: aujourd'hui)"),
            ("Date livraison prévue", "Format: AAAA-MM-JJ (optionnel)"),
            ("Campagne", "Période campagne: 2024-2025 (défaut: campagne courante)"),
            ("Certifications", "Noms des certifications séparés par virgule"),
            ("Taux droits export (%)", "Taux en pourcentage (défaut: config société)"),
            ("Notes", "Notes additionnelles (optionnel)"),
            ("", ""),
            ("IMPORTANT", "Les champs marqués * sont obligatoires"),
            ("Types de produit valides:", "cocoa_mass, cocoa_butter, cocoa_cake, cocoa_powder"),
        ]
        
        for row_idx, (col_a, col_b) in enumerate(instructions, start=1):
            ws_help[f"A{row_idx}"] = col_a
            ws_help[f"B{row_idx}"] = col_b
            if row_idx == 1:
                ws_help[f"A{row_idx}"].font = header_font
                ws_help[f"B{row_idx}"].font = header_font
                ws_help[f"A{row_idx}"].fill = header_fill
                ws_help[f"B{row_idx}"].fill = header_fill
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    # -------------------------------------------------------------------------
    # ACTION METHODS
    # -------------------------------------------------------------------------
    def action_download_template(self):
        """Download the Excel template."""
        self.ensure_one()
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_(
                "La bibliothèque 'openpyxl' n'est pas installée. "
                "Veuillez contacter l'administrateur pour l'installer."
            ))
        
        template_data = self._generate_template()
        
        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'modele_import_contrats.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(template_data),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_parse_file(self):
        """Parse the uploaded Excel file and show preview."""
        self.ensure_one()
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_(
                "La bibliothèque 'openpyxl' n'est pas installée. "
                "Veuillez contacter l'administrateur pour l'installer."
            ))
        
        if not self.file_data:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))
        
        if not self.file_name or not self.file_name.endswith('.xlsx'):
            raise UserError(_("Le fichier doit être au format Excel (.xlsx)."))
        
        # Parse file
        try:
            file_content = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            ws = wb.active
        except Exception as e:
            raise UserError(_("Erreur lors de la lecture du fichier: %s") % str(e))
        
        # Clear existing preview lines
        self.preview_line_ids.unlink()
        
        # Parse rows (skip header)
        lines_vals = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue
            
            line_vals = self._parse_row(row, row_idx)
            lines_vals.append((0, 0, line_vals))
        
        if not lines_vals:
            raise UserError(_("Le fichier ne contient aucune donnée à importer."))
        
        self.write({
            'preview_line_ids': lines_vals,
            'state': 'preview',
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'potting.import.contracts.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _parse_row(self, row, row_idx):
        """Parse a single row from the Excel file."""
        errors = []
        
        # Extract values
        contract_number = str(row[0]).strip() if row[0] else ''
        customer_name = str(row[1]).strip() if row[1] else ''
        product_type = str(row[2]).strip() if row[2] else ''
        tonnage_str = row[3]
        unit_price_str = row[4]
        currency_code = str(row[5]).strip() if row[5] else 'EUR'
        date_order_str = row[6]
        date_expected_str = row[7]
        campaign = str(row[8]).strip() if row[8] else ''
        certifications_str = str(row[9]).strip() if row[9] else ''
        export_duty_rate_str = row[10]
        notes = str(row[11]).strip() if len(row) > 11 and row[11] else ''
        
        # Validate contract number
        if not contract_number:
            errors.append(_("Numéro de contrat manquant"))
        else:
            existing = self.env['potting.customer.order'].search([
                ('contract_number', '=', contract_number)
            ], limit=1)
            if existing:
                errors.append(_("Contrat '%s' existe déjà") % contract_number)
        
        # Validate customer
        customer_id = False
        if not customer_name:
            errors.append(_("Nom du client manquant"))
        else:
            customer = self.env['res.partner'].search([
                ('name', 'ilike', customer_name),
                ('is_company', '=', True)
            ], limit=1)
            if customer:
                customer_id = customer.id
            else:
                errors.append(_("Client '%s' non trouvé") % customer_name)
        
        # Validate product type
        valid_product_types = ['cocoa_mass', 'cocoa_butter', 'cocoa_cake', 'cocoa_powder']
        if not product_type:
            errors.append(_("Type de produit manquant"))
        elif product_type not in valid_product_types:
            errors.append(_("Type de produit invalide: %s") % product_type)
        
        # Validate tonnage
        tonnage = 0.0
        if tonnage_str:
            try:
                tonnage = float(tonnage_str)
                if tonnage <= 0:
                    errors.append(_("Le tonnage doit être positif"))
            except (ValueError, TypeError):
                errors.append(_("Tonnage invalide: %s") % tonnage_str)
        else:
            errors.append(_("Tonnage manquant"))
        
        # Parse unit price
        unit_price = 0.0
        if unit_price_str:
            try:
                unit_price = float(unit_price_str)
            except (ValueError, TypeError):
                errors.append(_("Prix unitaire invalide: %s") % unit_price_str)
        
        # Validate currency
        currency_id = False
        if currency_code:
            currency = self.env['res.currency'].search([
                ('name', '=', currency_code.upper())
            ], limit=1)
            if currency:
                currency_id = currency.id
            else:
                errors.append(_("Devise '%s' non trouvée") % currency_code)
        
        # Parse dates
        date_order = False
        if date_order_str:
            date_order = self._parse_date(date_order_str)
            if not date_order:
                errors.append(_("Format de date de commande invalide"))
        
        date_expected = False
        if date_expected_str:
            date_expected = self._parse_date(date_expected_str)
            if not date_expected:
                errors.append(_("Format de date de livraison invalide"))
        
        # Parse export duty rate
        export_duty_rate = 14.6  # Default
        if export_duty_rate_str:
            try:
                export_duty_rate = float(export_duty_rate_str)
            except (ValueError, TypeError):
                pass
        
        # Find certifications
        certification_ids = []
        if certifications_str:
            cert_names = [c.strip() for c in certifications_str.split(',')]
            for cert_name in cert_names:
                if cert_name:
                    cert = self.env['potting.certification'].search([
                        ('name', 'ilike', cert_name)
                    ], limit=1)
                    if cert:
                        certification_ids.append(cert.id)
        
        return {
            'row_number': row_idx,
            'contract_number': contract_number,
            'customer_name': customer_name,
            'customer_id': customer_id,
            'product_type': product_type if product_type in valid_product_types else False,
            'tonnage': tonnage,
            'unit_price': unit_price,
            'currency_id': currency_id or self.env.company.currency_id.id,
            'date_order': date_order or fields.Date.context_today(self),
            'date_expected': date_expected,
            'campaign_period': campaign or self._get_default_campaign_period(),
            'certification_ids': [(6, 0, certification_ids)] if certification_ids else False,
            'export_duty_rate': export_duty_rate,
            'notes': notes,
            'error_message': '\n'.join(errors) if errors else False,
            'is_valid': len(errors) == 0,
        }

    def _parse_date(self, date_val):
        """Parse a date from various formats."""
        if isinstance(date_val, datetime):
            return date_val.date()
        if isinstance(date_val, str):
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                try:
                    return datetime.strptime(date_val, fmt).date()
                except ValueError:
                    continue
        return False

    def _get_default_campaign_period(self):
        """Get default campaign period."""
        today = fields.Date.context_today(self)
        year = today.year
        if today.month >= 10:
            return f"{year}-{year + 1}"
        else:
            return f"{year - 1}-{year}"

    def action_import(self):
        """Import valid contracts."""
        self.ensure_one()
        
        valid_lines = self.preview_line_ids.filtered('is_valid')
        if not valid_lines:
            raise UserError(_("Aucune ligne valide à importer."))
        
        created_orders = self.env['potting.customer.order']
        errors = []
        
        for line in valid_lines:
            try:
                order_vals = {
                    'contract_number': line.contract_number,
                    'customer_id': line.customer_id.id,
                    'product_type': line.product_type,
                    'contract_tonnage': line.tonnage,  # Le tonnage du fichier devient le tonnage du contrat
                    'date_order': line.date_order,
                    'date_expected': line.date_expected,
                    'campaign_period': line.campaign_period,
                    'currency_id': line.currency_id.id,
                    'unit_price': line.unit_price,
                    'export_duty_rate': line.export_duty_rate,
                    'note': line.notes,
                }
                
                # Create order
                order = self.env['potting.customer.order'].create(order_vals)
                
                # Add certifications
                if line.certification_ids:
                    order.certification_ids = line.certification_ids
                
                # Note: Les OT seront créés par l'utilisateur Shipping ultérieurement
                # Le contrat est créé avec le tonnage prévu (contract_tonnage)
                
                created_orders |= order
                
            except Exception as e:
                errors.append(_("Ligne %s: %s") % (line.row_number, str(e)))
        
        # Prepare result message
        result_msg = _("Import terminé.\n\n")
        result_msg += _("✅ %s contrat(s) créé(s) avec succès.\n") % len(created_orders)
        result_msg += _("ℹ️ Les OT peuvent maintenant être générés par l'équipe Shipping.\n")
        
        if errors:
            result_msg += _("\n⚠️ Erreurs:\n")
            result_msg += '\n'.join(errors)
        
        self.write({
            'state': 'done',
            'import_result': result_msg,
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'potting.import.contracts.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_view_orders(self):
        """View created orders."""
        return {
            'type': 'ir.actions.act_window',
            'name': _('Contrats importés'),
            'res_model': 'potting.customer.order',
            'view_mode': 'tree,form',
            'target': 'current',
            'domain': [('create_date', '>=', fields.Datetime.now().replace(hour=0, minute=0, second=0))],
        }


class PottingImportContractsWizardLine(models.TransientModel):
    """Ligne d'aperçu pour l'import de contrats."""
    
    _name = 'potting.import.contracts.wizard.line'
    _description = 'Ligne d\'import de contrat'

    wizard_id = fields.Many2one(
        'potting.import.contracts.wizard',
        string="Wizard",
        required=True,
        ondelete='cascade'
    )
    
    row_number = fields.Integer(string="Ligne")
    
    contract_number = fields.Char(string="N° Contrat")
    
    customer_name = fields.Char(string="Client (Excel)")
    
    customer_id = fields.Many2one(
        'res.partner',
        string="Client (Odoo)"
    )
    
    product_type = fields.Selection([
        ('cocoa_mass', 'Masse de cacao'),
        ('cocoa_butter', 'Beurre de cacao'),
        ('cocoa_cake', 'Cake (Tourteau) de cacao'),
        ('cocoa_powder', 'Poudre de cacao'),
    ], string="Type de produit")
    
    tonnage = fields.Float(
        string="Tonnage (T)",
        digits='Product Unit of Measure'
    )
    
    unit_price = fields.Float(string="Prix unitaire")
    
    currency_id = fields.Many2one(
        'res.currency',
        string="Devise"
    )
    
    date_order = fields.Date(string="Date commande")
    
    date_expected = fields.Date(string="Date livraison")
    
    campaign_period = fields.Char(string="Campagne")
    
    certification_ids = fields.Many2many(
        'potting.certification',
        string="Certifications"
    )
    
    export_duty_rate = fields.Float(string="Taux droits (%)")
    
    notes = fields.Text(string="Notes")
    
    is_valid = fields.Boolean(string="Valide", default=True)
    
    error_message = fields.Text(string="Erreurs")
